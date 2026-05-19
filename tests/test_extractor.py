"""Tests for content extractors."""

from __future__ import annotations

from pathlib import Path

from fileforge.extractor import extract_snippet


def test_extract_plain_text(tmp_dir: Path) -> None:
    """Plain .txt file returns its content as snippet."""
    f = tmp_dir / "HARNESS_doc.txt"
    f.write_text("This is a consulting proposal for a nonprofit.")
    snippet = extract_snippet(f, max_chars=2000)
    assert "consulting proposal" in snippet


def test_extract_markdown(tmp_dir: Path) -> None:
    """Markdown file returns text content."""
    f = tmp_dir / "HARNESS_readme.md"
    f.write_text("# Project\n\nThis project handles hydroponics automation.\n")
    snippet = extract_snippet(f, max_chars=2000)
    assert "hydroponics" in snippet


def test_extract_python(tmp_dir: Path) -> None:
    """Python source file returns its content."""
    f = tmp_dir / "HARNESS_main.py"
    f.write_text('def hello():\n    """Greet the user."""\n    return "hello"\n')
    snippet = extract_snippet(f, max_chars=2000)
    assert "def hello" in snippet


def test_extract_truncates_at_max_chars(tmp_dir: Path) -> None:
    """Snippet is capped at max_chars."""
    f = tmp_dir / "HARNESS_long.txt"
    f.write_text("x" * 5000)
    snippet = extract_snippet(f, max_chars=100)
    assert len(snippet) <= 100


def test_extract_unknown_binary_returns_none(tmp_dir: Path) -> None:
    """Unknown binary extension returns None (metadata-only record)."""
    f = tmp_dir / "HARNESS_blob.xyz"
    f.write_bytes(bytes(range(256)))
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is None


def test_extract_handles_encoding_detection(tmp_dir: Path) -> None:
    """Files with non-UTF-8 encoding are read without raising."""
    f = tmp_dir / "HARNESS_latin.txt"
    f.write_bytes("Résumé professionnel\n".encode("latin-1"))
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is not None
    assert len(snippet) > 0


def test_extract_docx_error_returns_none(tmp_dir: Path) -> None:
    """DOCX extraction on a non-DOCX file returns None gracefully."""
    f = tmp_dir / "HARNESS_fake.docx"
    f.write_bytes(b"not a real docx file")
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is None


def test_extract_whitespace_only_returns_none(tmp_dir: Path) -> None:
    """A file containing only whitespace returns None (no useful content)."""
    f = tmp_dir / "HARNESS_blank.txt"
    f.write_text("   \n\t\n   ")
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is None


# ── XLSX ──────────────────────────────────────────────────────────────────────


def _make_xlsx(path: Path, sheets: dict[str, list[list]]) -> None:
    """Helper: write an xlsx workbook with the given sheet data."""
    import openpyxl

    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        for row in rows:
            ws.append(row)
    wb.save(path)


def test_extract_xlsx_basic(tmp_dir: Path) -> None:
    """XLSX file with cell data returns text snippet."""
    f = tmp_dir / "HARNESS_budget.xlsx"
    _make_xlsx(
        f, {"Budget": [["Category", "Amount"], ["Salaries", 50000], ["Supplies", 1200]]}
    )
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is not None
    assert "Category" in snippet
    assert "Salaries" in snippet
    assert "50000" in snippet


def test_extract_xlsx_sheet_header(tmp_dir: Path) -> None:
    """Sheet name appears as a section header in the snippet."""
    f = tmp_dir / "HARNESS_report.xlsx"
    _make_xlsx(f, {"Q1 Report": [["Revenue", "900000"]]})
    snippet = extract_snippet(f, max_chars=2000)
    assert "[Q1 Report]" in snippet


def test_extract_xlsx_multi_sheet(tmp_dir: Path) -> None:
    """Multiple sheets are all included up to max_chars."""
    f = tmp_dir / "HARNESS_multi.xlsx"
    _make_xlsx(
        f,
        {
            "Sheet1": [["Alpha", "Beta"]],
            "Sheet2": [["Gamma", "Delta"]],
        },
    )
    snippet = extract_snippet(f, max_chars=2000)
    assert "Alpha" in snippet
    assert "Gamma" in snippet


def test_extract_xlsx_respects_max_chars(tmp_dir: Path) -> None:
    """Snippet is capped at max_chars for large workbooks."""
    f = tmp_dir / "HARNESS_large.xlsx"
    _make_xlsx(f, {"Data": [["x" * 100] for _ in range(50)]})
    snippet = extract_snippet(f, max_chars=200)
    assert snippet is not None
    assert len(snippet) <= 220  # slight slack for sheet header


def test_extract_xlsx_empty_workbook_returns_none(tmp_dir: Path) -> None:
    """An xlsx with no cell values returns None."""
    import openpyxl

    f = tmp_dir / "HARNESS_empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(f)
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is None


def test_extract_xlsx_skips_empty_cells(tmp_dir: Path) -> None:
    """Empty cells in rows are not emitted as blank separators."""
    f = tmp_dir / "HARNESS_sparse.xlsx"
    _make_xlsx(f, {"Sparse": [["Hello", None, None, "World"]]})
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is not None
    assert "Hello" in snippet
    assert "World" in snippet
    # None values should not appear literally
    assert "None" not in snippet


def test_extract_xlsx_corrupt_returns_none(tmp_dir: Path) -> None:
    """Corrupt/non-xlsx file with .xlsx extension returns None gracefully."""
    f = tmp_dir / "HARNESS_corrupt.xlsx"
    f.write_bytes(b"not a real xlsx file")
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is None


def test_extract_pdf_corrupt_returns_none(tmp_dir: Path) -> None:
    """PDF extraction on a corrupt (non-PDF) file returns None gracefully."""
    f = tmp_dir / "HARNESS_corrupt.pdf"
    f.write_bytes(b"this is not a real pdf file")
    snippet = extract_snippet(f, max_chars=2000)
    assert snippet is None


# ── Image / OCR ───────────────────────────────────────────────────────────────


def test_image_extensions_registered() -> None:
    """Image extensions are registered in the dispatcher."""
    from fileforge.extractor import _DISPATCH

    for ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp", ".gif"):
        assert ext in _DISPATCH, f"{ext} not registered"


def test_image_extract_without_pytesseract_returns_none(tmp_dir: Path) -> None:
    """extract() returns None gracefully when pytesseract is not installed."""
    from fileforge.extractor import image as _image_mod

    # If pytesseract genuinely isn't available, _AVAILABLE is False and
    # extract() returns None without touching the file.
    if _image_mod._AVAILABLE:
        # pytesseract IS installed in this environment — skip the unavailability test
        return

    f = tmp_dir / "HARNESS_scan.png"
    # Minimal 1×1 white PNG (valid file, but OCR won't run)
    from PIL import Image as PILImage

    PILImage.new("RGB", (1, 1), color=255).save(f)

    snippet = _image_mod.extract(f, max_chars=2000)
    assert snippet is None


def test_image_extract_via_mock(tmp_dir: Path) -> None:
    """extract() returns OCR text when pytesseract is available (mocked)."""
    from unittest.mock import MagicMock, patch

    from PIL import Image as PILImage

    f = tmp_dir / "HARNESS_ocr.png"
    PILImage.new("RGB", (100, 30), color=255).save(f)

    mock_pytesseract = MagicMock()
    mock_pytesseract.TesseractNotFoundError = Exception
    mock_pytesseract.image_to_string.return_value = "Invoice total: $1,234.56"

    with (
        patch("fileforge.extractor.image._AVAILABLE", True),
        patch("fileforge.extractor.image.pytesseract", mock_pytesseract, create=True),
    ):
        from fileforge.extractor import image as _image_mod

        snippet = _image_mod.extract(f, max_chars=2000)

    assert snippet == "Invoice total: $1,234.56"


def test_image_extract_via_mock_truncates(tmp_dir: Path) -> None:
    """extract() truncates OCR output at max_chars."""
    from unittest.mock import MagicMock, patch

    from PIL import Image as PILImage

    f = tmp_dir / "HARNESS_long_ocr.png"
    PILImage.new("RGB", (100, 30), color=255).save(f)

    mock_pytesseract = MagicMock()
    mock_pytesseract.TesseractNotFoundError = Exception
    mock_pytesseract.image_to_string.return_value = "A" * 5000

    with (
        patch("fileforge.extractor.image._AVAILABLE", True),
        patch("fileforge.extractor.image.pytesseract", mock_pytesseract, create=True),
    ):
        from fileforge.extractor import image as _image_mod

        snippet = _image_mod.extract(f, max_chars=100)

    assert len(snippet) == 100


def test_image_extract_empty_ocr_returns_none(tmp_dir: Path) -> None:
    """extract() returns None when OCR produces only whitespace."""
    from unittest.mock import MagicMock, patch

    from PIL import Image as PILImage

    f = tmp_dir / "HARNESS_blank_img.png"
    PILImage.new("RGB", (100, 30), color=255).save(f)

    mock_pytesseract = MagicMock()
    mock_pytesseract.TesseractNotFoundError = Exception
    mock_pytesseract.image_to_string.return_value = "   \n\n\t  "

    with (
        patch("fileforge.extractor.image._AVAILABLE", True),
        patch("fileforge.extractor.image.pytesseract", mock_pytesseract, create=True),
    ):
        from fileforge.extractor import image as _image_mod

        snippet = _image_mod.extract(f, max_chars=2000)

    assert snippet is None


def test_image_extract_tesseract_not_found_returns_none(tmp_dir: Path) -> None:
    """extract() returns None when the Tesseract binary is missing."""
    from unittest.mock import MagicMock, patch

    from PIL import Image as PILImage

    f = tmp_dir / "HARNESS_notesseract.png"
    PILImage.new("RGB", (100, 30), color=255).save(f)

    class _TesseractNotFoundError(Exception):
        pass

    mock_pytesseract = MagicMock()
    mock_pytesseract.TesseractNotFoundError = _TesseractNotFoundError
    mock_pytesseract.image_to_string.side_effect = _TesseractNotFoundError("not found")

    with (
        patch("fileforge.extractor.image._AVAILABLE", True),
        patch("fileforge.extractor.image.pytesseract", mock_pytesseract, create=True),
    ):
        from fileforge.extractor import image as _image_mod

        snippet = _image_mod.extract(f, max_chars=2000)

    assert snippet is None
